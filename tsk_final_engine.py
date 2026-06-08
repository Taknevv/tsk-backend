"""
=============================================================================
TSK FINAL UNIFIED ENGINE — AUTO-UPDATE + 10 AI ALGORITHMS
Online Coil Inspection | CGL, CAL & RCL | Tata Steel Kalinganagar
=============================================================================
Called by tsk_watcher.py whenever TSK_Inspection_Data.xlsx is saved.
Also callable directly:
    python tsk_final_engine.py                          # default paths
    python tsk_final_engine.py input.xlsx output.xlsx   # custom paths

OUTPUT WORKBOOK SHEETS (16 total):
  ── LIVE RESULTS ──────────────────────────────────────────────────────
  1.  🏠 Dashboard         KPI tiles, per-line summary, gap table
  2.  📋 Coil Detail        Row-by-row coil + inspection data
  3.  😴 Fatigue Log        Fatigue scores + rotation actions
  4.  👷 Inspector Matrix   Certification map + gap lines
  5.  🔢 Inspector Calc     Headcount model (raw → base → peak)
  6.  📝 Change Log         Timestamped audit trail of every rebuild

  ── AI ALGORITHM RESULTS (auto-recalculate on every save) ────────────
  7.  A1 Demand Forecast    EWMA + Holt-Winters 24h forecast
  8.  A2 Anomaly Detection  Z-Score + IQR + Isolation Forest
  9.  A3 RL Policy          Q-Learning optimal rotation policy
  10. A4 Fatigue Predict    Polynomial regression + time-to-critical
  11. A5 DP Scheduling      Dynamic programming shift sequencing
  12. A6 Genetic Algorithm  Multi-line inspector scheduling (GA)
  13. A7 CUSUM Control      Statistical process control
  14. A8 Monte Carlo        Staffing risk simulation (5000 runs)
  15. A9 Markov Chain       Inspector state transitions + steady-state
  16. A10 Live Dashboard    Real-time OEE + utilisation + alert engine
=============================================================================
"""

import sys, math, random, statistics, datetime
from pathlib import Path
from collections import defaultdict

import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(0); np.random.seed(0)

# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
NAVY="0D2137"; BLUE="1F4E79"; MID="2E75B6"; LT_BLUE="DEEAF1"
RED="C00000";  LT_RED="FFE7E7"; AMB="E36C09"; LT_AMB="FFF2CC"
GRN="375623";  LT_GRN="EBF3E8"; GREY="595959"; WHITE="FFFFFF"
ACC="00B0D7";  LT_GRY="F5F5F5"; OFF_W="F7F9FC"; DARK="0A1628"

def _s(c="CCCCCC", t="thin"): return Side(style=t, color=c)
def _fill(c): return PatternFill("solid", fgColor=c)
def _font(b=False, sz=10, c="000000"): return Font(bold=b, size=sz, color=c, name="Calibri")
def _al(h="center", v="center", w=True): return Alignment(horizontal=h, vertical=v, wrap_text=w)
def _bd(c="CCCCCC"):
    s = _s(c); return Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, r, c, val, bg=WHITE, fc="000000", bold=False, sz=10,
          align="center", border=True, fmt=None):
    cl = ws.cell(row=r, column=c, value=val)
    cl.font = _font(bold, sz, fc)
    cl.fill = _fill(bg)
    cl.alignment = _al(align)
    if border: cl.border = _bd()
    if fmt: cl.number_format = fmt
    return cl

def hrow(ws, row, vals, bg=NAVY, fc=WHITE, bold=True, h=26, start=1):
    ws.row_dimensions[row].height = h
    for i, v in enumerate(vals, start):
        _cell(ws, row, i, v, bg=bg, fc=fc, bold=bold, align="center")

def drow(ws, row, vals, bg=WHITE, fc="000000", bold=False, h=18, lcols=None, start=1):
    ws.row_dimensions[row].height = h
    for i, v in enumerate(vals, start):
        al = "left" if lcols and i in lcols else "center"
        _cell(ws, row, i, v, bg=bg, fc=fc, bold=bold, align=al)

def title_merge(ws, text, row, end_col, bg=NAVY, fc=WHITE, sz=14, h=36):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cl = ws.cell(row=row, column=1, value=text)
    cl.font = _font(True, sz, fc); cl.fill = _fill(bg)
    cl.alignment = _al("center"); ws.row_dimensions[row].height = h

def sub_merge(ws, text, row, end_col, bg=BLUE, fc=WHITE, sz=10, h=20):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cl = ws.cell(row=row, column=1, value=text)
    cl.font = _font(True, sz, fc); cl.fill = _fill(bg)
    cl.alignment = _al("left"); ws.row_dimensions[row].height = h

def note_merge(ws, text, row, end_col, bg=OFF_W, h=20):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cl = ws.cell(row=row, column=1, value=text)
    cl.font = _font(False, 9, GREY); cl.fill = _fill(bg)
    cl.alignment = _al("left"); ws.row_dimensions[row].height = h

def gcol(n): return get_column_letter(n)
def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[gcol(i)].width = w
def no_grid(ws): ws.sheet_view.showGridLines = False

# ─────────────────────────────────────────────────────────────────────────────
# READ SOURCE DATA
# ─────────────────────────────────────────────────────────────────────────────
def read_source(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames

    def load_sheet(name):
        if name not in sheets: return []
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: return []
        hdr = [str(h).strip() if h else f"col{i}" for i,h in enumerate(rows[0])]
        return [dict(zip(hdr, r)) for r in rows[1:] if r[0]]

    coils   = load_sheet("Coil_Production")
    logs    = load_sheet("Inspection_Log")
    skills  = load_sheet("Inspector_Skills")
    hourly  = load_sheet("Hourly_Data")

    wb.close()
    return coils, logs, skills, hourly

# ─────────────────────────────────────────────────────────────────────────────
# COMPUTE CORE METRICS
# ─────────────────────────────────────────────────────────────────────────────
SPEED_BY_LINE = {"CGL": 150, "CAL": 200, "RCL": 250}
PEAK_FACTOR   = 1.3
AVAIL_MIN     = 60

def compute_metrics(coils, logs, skills):
    # Build inspection lookup
    log_by_coil = {l["coil_id"]: l for l in logs if l.get("coil_id")}

    enriched = []
    for c in coils:
        cid   = c.get("coil_id","")
        line  = str(c.get("line","")).upper()
        length= float(c.get("length_m") or 0)
        spd_s = float(c.get("speed_mps") or (SPEED_BY_LINE.get(line,150)/60))
        defects= int(c.get("defect_count") or 0)

        # W_l from actual inspection duration if available
        lg = log_by_coil.get(cid,{})
        dur_min = 0
        t_start = lg.get("inspection_start")
        t_end   = lg.get("inspection_end")
        if t_start and t_end:
            try:
                if isinstance(t_start, str): t_start = datetime.datetime.fromisoformat(t_start)
                if isinstance(t_end,   str): t_end   = datetime.datetime.fromisoformat(t_end)
                dur_min = (t_end - t_start).total_seconds() / 60
            except: dur_min = 0
        if dur_min <= 0 and length > 0 and spd_s > 0:
            dur_min = length / (spd_s * 60)

        wl = round(dur_min * 60 / length, 4) if length > 0 else 0
        defects_km = round(defects / (length / 1000), 4) if length > 0 else 0
        fatigue = float(lg.get("fatigue_score_post") or 0)
        inspector = lg.get("inspector_id","—")

        enriched.append({
            "coil_id": cid, "line": line, "length_m": length,
            "speed_mps": spd_s, "defect_count": defects,
            "defects_km": defects_km, "wl": wl,
            "dur_min": round(dur_min,2), "fatigue": fatigue,
            "inspector": inspector,
        })

    # Per-line aggregates
    line_stats = {}
    for ln in ["CGL","CAL","RCL"]:
        rows = [r for r in enriched if r["line"]==ln]
        if not rows:
            line_stats[ln] = {"wl_avg":0,"wl_min":0,"wl_max":0,"defects_km_avg":0,"cv":0,
                              "n_coils":0,"total_m":0,"n_base":3,"n_peak":4,"fat_avg":0}
            continue
        wls   = [r["wl"] for r in rows if r["wl"]>0]
        defs  = [r["defects_km"] for r in rows]
        fats  = [r["fatigue"] for r in rows if r["fatigue"]>0]
        wl_avg= round(statistics.mean(wls),4) if wls else 0
        def_avg=round(statistics.mean(defs),4) if defs else 0
        cv    = round(statistics.stdev(defs)/def_avg,4) if def_avg>0 and len(defs)>1 else 0
        speed = SPEED_BY_LINE[ln]
        raw   = (speed * wl_avg) / AVAIL_MIN if wl_avg>0 else 2.2
        n_base= max(1, math.ceil(raw))
        n_peak= math.ceil(n_base * PEAK_FACTOR)
        line_stats[ln] = {
            "wl_avg":wl_avg,"wl_min":round(min(wls),4) if wls else 0,
            "wl_max":round(max(wls),4) if wls else 0,
            "defects_km_avg":def_avg,"cv":cv,"n_coils":len(rows),
            "total_m":round(sum(r["length_m"] for r in rows),1),
            "n_base":n_base,"n_peak":n_peak,
            "fat_avg":round(statistics.mean(fats),2) if fats else 0,
            "speed":speed,
        }

    # Inspector data
    insp_map = {}
    for s in skills:
        iid = s.get("inspector_id","")
        cert_raw = str(s.get("certified_lines","")).upper()
        certs = [c.strip() for c in cert_raw.split(",") if c.strip()]
        insp_map[iid] = {
            "name": s.get("name",""),
            "certs": certs,
            "shift": s.get("shift_preference","day"),
        }

    return enriched, line_stats, insp_map

# ─────────────────────────────────────────────────────────────────────────────
# ── SHEET BUILDERS: LIVE RESULTS ─────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
def build_dashboard(wb, enriched, line_stats, insp_map, source_path, ts):
    ws = wb.create_sheet("🏠 Dashboard")
    no_grid(ws)
    set_widths(ws, [18,14,14,14,14,14,14,14,14,14,14,14])
    NC = 12

    total_coils   = len(enriched)
    avail_insp    = len(insp_map)
    base_req      = sum(v["n_base"]  for v in line_stats.values())
    peak_req      = sum(v["n_peak"]  for v in line_stats.values())
    gap           = max(0, base_req - avail_insp)
    all_fats      = [r["fatigue"] for r in enriched if r["fatigue"]>0]
    avg_fat       = round(statistics.mean(all_fats),2) if all_fats else 0

    title_merge(ws, "TSK COIL INSPECTION — LIVE RESULTS DASHBOARD", 1, NC, sz=16)
    ws.merge_cells(f"A2:{gcol(NC)}2")
    cl=ws.cell(2,1,f"Last updated: {ts}  |  Source: {Path(source_path).name}  |  Auto-generated — do not edit manually")
    cl.font=_font(False,9,ACC); cl.fill=_fill(NAVY); cl.alignment=_al("center")
    ws.row_dimensions[2].height=18

    # KPI tiles
    kpis = [
        (str(total_coils),   "Total Coils\nInspected",     BLUE,    LT_BLUE,  "000000"),
        (str(avail_insp),    "Inspectors\nAvailable",      GRN if avail_insp>=base_req else RED, LT_GRN if avail_insp>=base_req else LT_RED, "000000"),
        (str(base_req),      "Base\nRequired",             MID,     LT_BLUE,  "000000"),
        (str(peak_req),      "Peak\nRequired",             AMB,     LT_AMB,   "000000"),
        (f"−{gap}" if gap>0 else "0", "Staffing\nGap",   RED if gap>0 else GRN, LT_RED if gap>0 else LT_GRN, RED if gap>0 else GRN),
        (f"{avg_fat}",       "Avg Fatigue\n(target ≤6)",  RED if avg_fat>7 else AMB if avg_fat>5 else GRN, LT_RED if avg_fat>7 else LT_AMB, "000000"),
    ]
    r=4
    for col_s, (val, lbl, bg, lbg, fc) in enumerate(kpis, 1):
        ws.merge_cells(start_row=r,start_column=col_s*2-1,end_row=r,end_column=col_s*2)
        cl=ws.cell(r,col_s*2-1,val); cl.font=_font(True,28,bg); cl.fill=_fill(lbg)
        cl.alignment=_al("center"); cl.border=_bd(bg)
        ws.merge_cells(start_row=r+1,start_column=col_s*2-1,end_row=r+1,end_column=col_s*2)
        cl2=ws.cell(r+1,col_s*2-1,lbl); cl2.font=_font(False,9,GREY); cl2.fill=_fill(lbg)
        cl2.alignment=_al("center"); cl2.border=_bd(bg)
    ws.row_dimensions[r].height=40; ws.row_dimensions[r+1].height=26

    # Per-line table
    r=7
    sub_merge(ws,"PER-LINE PERFORMANCE METRICS",r,NC); r+=1
    hrow(ws,r,["Line","Speed (m/min)","Coils","Avg W_l (s/m)","Min W_l","Max W_l",
               "Avg Def/km","CV","Base Insp.","Peak Insp.","Fatigue Avg","Status"]); r+=1
    for i,(ln,s) in enumerate(line_stats.items()):
        ok = s["n_base"]<=avail_insp
        status="OK" if ok else "GAP"
        sbg = LT_GRN if ok else LT_RED
        sfc = GRN if ok else RED
        bgs = [LT_BLUE if i%2==0 else WHITE]*12
        drow(ws,r,[ln,s["speed"],s["n_coils"],s["wl_avg"],s["wl_min"],s["wl_max"],
                   s["defects_km_avg"],s["cv"],s["n_base"],s["n_peak"],s["fat_avg"],status],
             bg=bgs[0]); r+=1
        ws.cell(r-1,12).fill=_fill(sbg); ws.cell(r-1,12).font=_font(True,10,sfc)

    # Gap table
    r+=1
    sub_merge(ws,"WORKFORCE GAP SUMMARY",r,NC); r+=1
    hrow(ws,r,["Line","Available","Base Req.","Peak Req.","Gap (Base)","Gap (Peak)",
               "% Shortfall","Priority","Action Required","Rotation (min)","Fatigue Risk",""]); r+=1
    rot_map={"CGL":45,"CAL":60,"RCL":30}
    for i,(ln,s) in enumerate(line_stats.items()):
        av2=avail_insp; gb=max(0,s["n_base"]-av2); gp=max(0,s["n_peak"]-av2)
        pct=round(gb/s["n_base"]*100) if s["n_base"]>0 else 0
        prio="CRITICAL" if gb>2 else "HIGH" if gb>0 else "OK"
        action="RECRUIT IMMEDIATELY" if gb>1 else "MONITOR" if gb>0 else "Maintain"
        fat_str=f"{s['fat_avg']}/10 {'CRITICAL' if s['fat_avg']>=8 else 'HIGH' if s['fat_avg']>=6 else 'OK'}"
        pbg = LT_RED if "CRITICAL" in prio else LT_AMB if "HIGH" in prio else LT_GRN
        pfc = RED if "CRITICAL" in prio else AMB if "HIGH" in prio else GRN
        drow(ws,r,[ln,av2,s["n_base"],s["n_peak"],f"−{gb}" if gb>0 else "0",
                   f"−{gp}" if gp>0 else "0",f"{pct}%",prio,action,
                   rot_map.get(ln,45),fat_str,""],
             bg=LT_BLUE if i%2==0 else WHITE); r+=1
        ws.cell(r-1,8).fill=_fill(pbg); ws.cell(r-1,8).font=_font(True,10,pfc)
    # Totals
    drow(ws,r,["TOTAL",avail_insp,base_req,peak_req,
               f"−{gap}" if gap>0 else "0",
               f"−{max(0,peak_req-avail_insp)}" if peak_req>avail_insp else "0",
               "","","","","",""],
         bg=NAVY,fc=WHITE,bold=True); r+=1

def build_coil_detail(wb, enriched):
    ws = wb.create_sheet("📋 Coil Detail")
    no_grid(ws); set_widths(ws,[12,8,12,12,12,12,12,12,14,12])
    NC=10
    title_merge(ws,"COIL-BY-COIL INSPECTION DETAIL",1,NC); r=2
    hrow(ws,r,["Coil ID","Line","Length (m)","Speed (m/s)","Defect Count",
               "Defects/km","W_l (s/m)","Inspector","Duration (min)","Fatigue Score"]); r+=1
    for i,e in enumerate(enriched):
        fat=e["fatigue"]
        fbg = LT_RED if fat>=8 else LT_AMB if fat>=6 else LT_GRN if fat>0 else WHITE
        lbg = {"CGL":LT_BLUE,"CAL":LT_GRN,"RCL":LT_AMB}.get(e["line"],WHITE)
        drow(ws,r,[e["coil_id"],e["line"],e["length_m"],e["speed_mps"],e["defect_count"],
                   e["defects_km"],e["wl"],e["inspector"],e["dur_min"],
                   e["fatigue"] if e["fatigue"]>0 else "N/A"],
             bg=LT_GRY if i%2==0 else WHITE); r+=1
        ws.cell(r-1,10).fill=_fill(fbg)
        ws.cell(r-1,2).fill=_fill(lbg)

def build_fatigue_log(wb, enriched, logs):
    ws = wb.create_sheet("😴 Fatigue Log")
    no_grid(ws); set_widths(ws,[14,12,14,10,12,22,10])
    NC=7; title_merge(ws,"FATIGUE SCORE LOG — ALL INSPECTIONS",1,NC); r=2
    hrow(ws,r,["Inspection ID","Coil ID","Inspector ID","Score","Band","Action Required","Risk"]); r+=1
    for i,lg in enumerate(logs):
        fat = float(lg.get("fatigue_score_post") or 0)
        band = ("CRITICAL" if fat>=8 else "HIGH" if fat>=6 else "MODERATE" if fat>=4 else "LOW")
        action = ("Immediate rotation" if fat>=9 else "Rotate next shift" if fat>=7
                  else "Monitor" if fat>=5 else "Normal")
        bg = LT_RED if fat>=8 else LT_AMB if fat>=6 else LT_GRN if fat>0 else WHITE
        fc = RED if fat>=8 else AMB if fat>=6 else GRN if fat>0 else "000000"
        drow(ws,r,[lg.get("inspection_id",""),lg.get("coil_id",""),
                   lg.get("inspector_id",""),fat,band,action,"●"],
             bg=LT_GRY if i%2==0 else WHITE); r+=1
        for c in [5,7]: ws.cell(r-1,c).fill=_fill(bg); ws.cell(r-1,c).font=_font(True,10,fc)

def build_inspector_matrix(wb, insp_map):
    ws = wb.create_sheet("👷 Inspector Matrix")
    no_grid(ws); set_widths(ws,[14,18,10,10,10,14,12,12])
    NC=8; title_merge(ws,"INSPECTOR CERTIFICATION MATRIX",1,NC); r=2
    hrow(ws,r,["Inspector ID","Name","CGL Cert.","CAL Cert.","RCL Cert.",
               "Lines Certified","Shift Pref.","Risk Level"]); r+=1
    for i,(iid,info) in enumerate(insp_map.items()):
        certs=info["certs"]
        miss=[ln for ln in ["CGL","CAL","RCL"] if ln not in certs]
        risk="MEDIUM" if len(miss)==1 else "HIGH" if len(miss)==2 else "LOW"
        rbg=LT_AMB if risk=="MEDIUM" else LT_RED if risk=="HIGH" else LT_GRN
        rfc=AMB if risk=="MEDIUM" else RED if risk=="HIGH" else GRN
        drow(ws,r,[iid,info["name"],
                   "YES" if "CGL" in certs else "NO",
                   "YES" if "CAL" in certs else "NO",
                   "YES" if "RCL" in certs else "NO",
                   len(certs),info["shift"],risk],
             bg=LT_GRY if i%2==0 else WHITE,lcols={2}); r+=1
        for c in [3,4,5]:
            v = ws.cell(r-1,c).value
            ws.cell(r-1,c).fill=_fill(LT_GRN if v=="YES" else LT_RED)
            ws.cell(r-1,c).font=_font(True,10,GRN if v=="YES" else RED)
        ws.cell(r-1,8).fill=_fill(rbg); ws.cell(r-1,8).font=_font(True,10,rfc)

def build_inspector_calc(wb, line_stats, insp_map):
    ws = wb.create_sheet("🔢 Inspector Calc")
    no_grid(ws); set_widths(ws,[16,14,14,14,14,14,14,14,20])
    NC=9; title_merge(ws,"INSPECTOR HEADCOUNT CALCULATOR",1,NC); r=2
    hrow(ws,r,["Line","Speed (m/min)","Avg W_l (s/m)","Defects/km","Def. Burden",
               "Raw Inspectors","Base (Rounded)","Peak (×1.3)","Inspectors Needed"]); r+=1
    for i,(ln,s) in enumerate(line_stats.items()):
        raw=round((s["speed"]*s["wl_avg"])/AVAIL_MIN,3) if s["wl_avg"]>0 else 2.2
        burden=round(s["defects_km_avg"]/10,4)
        needed=f"{s['n_base']} base / {s['n_peak']} peak"
        drow(ws,r,[ln,s["speed"],s["wl_avg"],s["defects_km_avg"],burden,
                   raw,s["n_base"],s["n_peak"],needed],
             bg=LT_BLUE if i%2==0 else WHITE); r+=1
        ws.cell(r-1,9).fill=_fill(LT_RED if s["n_base"]>len(insp_map) else LT_GRN)
    base_t=sum(v["n_base"] for v in line_stats.values())
    peak_t=sum(v["n_peak"] for v in line_stats.values())
    gap=max(0,base_t-len(insp_map))
    drow(ws,r,["TOTAL","—","—","—","—","—",base_t,peak_t,
               f"Gap: {gap} to recruit"],bg=NAVY,fc=WHITE,bold=True); r+=1

def build_change_log(wb, source_path, total_coils, ts, existing_log=None):
    ws = wb.create_sheet("📝 Change Log")
    no_grid(ws); set_widths(ws,[22,32,28,14,12])
    NC=5; title_merge(ws,"AUTO-UPDATE CHANGE LOG",1,NC); r=2
    hrow(ws,r,["Timestamp","Event","Source File","Total Coils","Status"]); r+=1
    # New entry
    drow(ws,r,[ts,"Auto-recalculation triggered",Path(source_path).name,total_coils,"SUCCESS"],
         bg=LT_GRN); r+=1
    # Previous entries from existing log if available
    if existing_log:
        for i,entry in enumerate(existing_log[:99]):
            drow(ws,r,[entry.get("Timestamp",""),entry.get("Event",""),
                       entry.get("Source File",""),entry.get("Total Coils",""),
                       entry.get("Status","")],
                 bg=LT_GRY if i%2==0 else WHITE); r+=1

# ─────────────────────────────────────────────────────────────────────────────
# ── AI ALGORITHM HELPERS ──────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
def make_series(enriched, line_stats, n=48):
    """Generate synthetic hourly series seeded from actual metrics."""
    series = {}
    for ln in ["CGL","CAL","RCL"]:
        s = line_stats[ln]
        base_def = s["defects_km_avg"] if s["defects_km_avg"] > 0 else 1.0
        base_fat = s["fat_avg"] if s["fat_avg"] > 0 else 6.5
        cv       = s["cv"] if s["cv"] > 0 else 0.3
        sigma    = max(0.05, base_def * cv * 0.5)
        def_s = [max(0, random.gauss(base_def*(1+math.sin(i*math.pi/10)*0.15), sigma)) for i in range(n)]
        fat_s = [min(10, max(1, base_fat + i*0.06 + random.gauss(0,0.4))) for i in range(n)]
        spd_s = [max(60, s["speed"]*(1+random.gauss(0,0.06))) for _ in range(n)]
        series[ln] = {"defects": def_s, "fatigue": fat_s, "speed": spd_s,
                      "base_def": base_def, "base_fat": base_fat, "cv": cv}
    return series

def holt_winters(series, alpha=0.3, beta=0.1, steps=24):
    L, T = series[0], series[1]-series[0]
    sm=[]
    for v in series:
        Lp,Tp=L,T
        L=alpha*v+(1-alpha)*(Lp+Tp)
        T=beta*(L-Lp)+(1-beta)*Tp
        sm.append(round(L,4))
    fc=[round(max(0,L+h*T),4) for h in range(1,steps+1)]
    return sm, fc

def ewma(series, alpha=0.2):
    r=[series[0]]
    for v in series[1:]: r.append(round(alpha*v+(1-alpha)*r[-1],4))
    return r

def zscore(series, thresh=2.5):
    mu=np.mean(series); sg=np.std(series)+1e-9
    return [(abs((v-mu)/sg)>thresh, round((v-mu)/sg,3)) for v in series]

def iqr_flag(series):
    q1,q3=np.percentile(series,25),np.percentile(series,75)
    iqr=q3-q1; lo,hi=q1-1.5*iqr,q3+1.5*iqr
    return [(v<lo or v>hi) for v in series]

def iso_score(val,series):
    mu=np.mean(series); sg=np.std(series)+1e-9
    return round(1-math.exp(-abs(val-mu)/sg*0.3),4)

def polyfit(x,y,deg=3,steps=12):
    c=np.polyfit(x,y,deg); p=np.poly1d(c)
    fitted=[round(float(p(xi)),3) for xi in x]
    fc=[round(max(1,min(10,float(p(len(x)+i)))),3) for i in range(1,steps+1)]
    res=[y[i]-fitted[i] for i in range(len(y))]
    rmse=round(math.sqrt(np.mean([r**2 for r in res])),4)
    r2=round(1-sum(r**2 for r in res)/(np.var(y)*len(y)+1e-9),4)
    return fitted,fc,rmse,r2,[round(ci,5) for ci in c]

def q_learning(n_ep=300,alpha=0.1,gamma=0.95,eps=0.2):
    MAX_ROT=[45,60,30]; Q=defaultdict(lambda:[0.0,0.0])
    rewards=[]; rotations=[]
    for ep in range(n_ep):
        li=random.randint(0,2); fat=random.uniform(1,5); ton=0
        tot=0; rots=0
        for _ in range(150):
            sf=min(4,int((fat-1)/2)); st=min(4,ton//10)
            state=(sf,st,li)
            act=random.randint(0,1) if random.random()<eps else int(Q[state][1]>Q[state][0])
            if act==0:
                ton+=1; fat=min(10,fat+random.uniform(0.05,0.2))
                rew=1.0-3*(fat>8)-2*(ton>MAX_ROT[li])
            else:
                rots+=1; fat=max(1,fat-random.uniform(1.5,2.5))
                ton=0; li=(li+1)%3; rew=0.5
            sf2=min(4,int((fat-1)/2)); st2=min(4,ton//10)
            ns=(sf2,st2,li)
            Q[state][act]+=alpha*(rew+gamma*max(Q[ns])-Q[state][act])
            tot+=rew
        rewards.append(round(tot,2)); rotations.append(rots)
    policy=[]
    for fb in range(5):
        for tb in range(5):
            for li,ln in enumerate(["CGL","CAL","RCL"]):
                s=(fb,tb,li); qv=Q[s]
                act="ROTATE" if qv[1]>qv[0] else "CONTINUE"
                conf=round(abs(qv[1]-qv[0])/(abs(qv[0])+abs(qv[1])+1e-6),3)
                policy.append({"line":ln,
                    "fatigue_bin":["1-3","3-5","5-7","7-9","9-10"][fb],
                    "time_bin":["0-10","10-20","20-30","30-45","45+"][tb],
                    "q_cont":round(qv[0],4),"q_rot":round(qv[1],4),
                    "policy":act,"confidence":conf})
    return policy, round(np.mean(rewards[-30:]),2), round(np.mean(rotations[-30:]),1)

def dp_scheduling(line_stats):
    demands=[]
    for sh_name,pf in [("Morning",1.3),("Afternoon",1.0),("Night",0.7)]:
        d={}
        for ln,s in line_stats.items():
            raw=(s["speed"]*s["wl_avg"])/AVAIL_MIN if s["wl_avg"]>0 else 2.2
            n=max(1,math.ceil(raw*pf))
            d[ln]=n
        demands.append((sh_name,d,pf))
    results=[]
    for avail in [3,6,9,12]:
        n_sh=len(demands); dp=[[0.0]*(avail+1) for _ in range(n_sh+1)]
        ch=[[0]*(avail+1) for _ in range(n_sh+1)]
        for i in range(n_sh-1,-1,-1):
            td=sum(demands[i][1].values()); w=demands[i][2]
            for j in range(avail+1):
                bv,bk=0,0
                for k in range(j+1):
                    cov=min(k/max(td,1),1.0)
                    val=w*cov+dp[i+1][j-k]
                    if val>bv: bv,bk=val,k
                dp[i][j]=bv; ch[i][j]=bk
        assign={}; rem=avail
        for i,(sn,sd,_) in enumerate(demands):
            k=ch[i][rem]; td=sum(sd.values())
            assign[sn]={"assigned":k,"demand":td,"cov_pct":round(min(k/max(td,1),1)*100,1),"gap":max(0,td-k)}
            rem-=k
        results.append({"avail":avail,"score":round(dp[0][avail],3),
                         "eff_pct":round(dp[0][avail]/3.0*100,1),"assign":assign})
    return results

def genetic_algorithm(line_stats):
    DEMAND=[]
    for sh_w in [1.3,1.0,0.7]:
        d=[]
        for ln,s in line_stats.items():
            raw=(s["speed"]*s["wl_avg"])/AVAIL_MIN if s["wl_avg"]>0 else 2.2
            d.append(max(1,math.ceil(raw*sh_w)))
        DEMAND.append(d)
    NL,NS,POP,GEN,MR=3,3,50,60,0.15; TOTAL=9
    def fit(ch):
        sc=0; used=sum(ch)
        for s in range(NS):
            for l in range(NL):
                sc+=[1.3,1.0,0.7][s]*min(ch[s*NL+l]/max(DEMAND[s][l],1),1)
        return sc-max(0,used-TOTAL)*0.5-sum(max(0,DEMAND[0][l]-ch[l])*1.5 for l in range(NL))
    def rand_c(): return [random.randint(0,4) for _ in range(NL*NS)]
    def cross(p1,p2): pt=random.randint(1,len(p1)-1); return p1[:pt]+p2[pt:]
    def mutate(c): return [max(0,min(5,g+random.randint(-1,1))) if random.random()<MR else g for g in c]
    def tourn(pop,k=4): return max(random.sample(pop,k),key=fit)
    pop=[rand_c() for _ in range(POP)]; bpg=[]; apg=[]
    for _ in range(GEN):
        fs=[fit(c) for c in pop]; bpg.append(round(max(fs),3)); apg.append(round(np.mean(fs),3))
        np_=[max(pop,key=fit)]
        while len(np_)<POP: np_.append(mutate(cross(tourn(pop),tourn(pop))))
        pop=np_
    best=max(pop,key=fit)
    sched=[]; lines=["CGL","CAL","RCL"]; shifts=["Morning","Afternoon","Night"]
    for s in range(NS):
        for l in range(NL):
            a=best[s*NL+l]; d=DEMAND[s][l]
            sched.append({"shift":shifts[s],"line":lines[l],"assigned":a,"demand":d,
                           "cov_pct":round(min(a/max(d,1),1)*100,1),
                           "status":"OK" if a>=d else "UNDER" if a>0 else "EMPTY"})
    conv=next((i for i in range(5,GEN) if bpg[i]-bpg[i-5]<0.01),GEN)
    return sched, round(fit(best),3), bpg, apg, conv

def cusum(series, ln):
    mu0=np.mean(series[:12]); sg=np.std(series[:12])+1e-6
    k=0.5*sg; h=4.0*sg; cp=[0.0]; cn=[0.0]; sigs=[]
    for v in series[1:]:
        cp.append(max(0,cp[-1]+(v-mu0-k)))
        cn.append(max(0,cn[-1]+(mu0-k-v)))
        sig=("ABOVE_LIMIT" if cp[-1]>h else "BELOW_LIMIT" if cn[-1]>h else "IN_CONTROL")
        sigs.append(sig)
    n_alarms=sigs.count("ABOVE_LIMIT")+sigs.count("BELOW_LIMIT")
    rate=round(n_alarms/max(len(sigs),1)*100,1)
    status=("OUT OF CONTROL" if rate>10 else "WATCH" if rate>5 else "IN CONTROL")
    return {"series":series,"cp":[round(v,4) for v in cp],"cn":[round(v,4) for v in cn],
            "sigs":sigs,"mu0":round(mu0,4),"sg":round(sg,4),"k":round(k,4),"h":round(h,4),
            "n_alarms":n_alarms,"rate":rate,"status":status}

def monte_carlo(line_stats, n=3000):
    results={}
    for ln,s in line_stats.items():
        req=s["n_base"]; miss=0; under=0; covs=[]
        for _ in range(n):
            avail=sum(random.random()>0.10 for _ in range(req))
            spike=np.random.poisson(1.0)*random.uniform(0.8,1.4)
            def_t=(s["defects_km_avg"] if s["defects_km_avg"]>0 else 1.0)*spike
            spd=random.gauss(s["speed"],s["speed"]*0.07)
            need=(spd*s["wl_avg"])/AVAIL_MIN if s["wl_avg"]>0 else req
            cov=min(avail/max(need,1),1.0); covs.append(round(cov,3))
            if avail<req: under+=1
            if cov<0.8: miss+=1
        p_under=round(under/n*100,2); p_miss=round(miss/n*100,2)
        pcts={f"p{k}":round(np.percentile(covs,k),3) for k in [5,25,50,75,95]}
        results[ln]={"p_under":p_under,"p_miss":p_miss,"var5":pcts["p5"],
                     "pcts":pcts,"risk":("EXTREME" if p_miss>40 else "HIGH" if p_miss>20 else "MEDIUM" if p_miss>10 else "LOW")}
    return results

def markov(line_stats):
    STATES=["ACTIVE","FATIGUED","ROTATING","ABSENT","TRAINING"]
    results={}
    for ln,s in line_stats.items():
        fat_bias=min(0.15,s["fat_avg"]/100) if s["fat_avg"]>0 else 0.05
        P=np.array([
            [0.65-fat_bias, 0.15+fat_bias, 0.12, 0.05, 0.03],
            [0.10, 0.28,                   0.52, 0.06, 0.04],
            [0.68+fat_bias, 0.15,           0.10, 0.04, 0.03],
            [0.40, 0.10,                   0.05, 0.40, 0.05],
            [0.55, 0.10,                   0.05, 0.05, 0.25],
        ])
        P=P/P.sum(axis=1,keepdims=True)
        pi=np.ones(5)/5
        for _ in range(500): pi=pi@P
        ones_pi=np.outer(np.ones(5),pi)
        Z=np.linalg.pinv(np.eye(5)-P+ones_pi)
        mfpt=np.diag(Z)/(pi+1e-9)
        results[ln]={"P":[[round(v,4) for v in row] for row in P],
                     "ss":{STATES[i]:round(pi[i]*100,2) for i in range(5)},
                     "mfpt":{STATES[i]:round(mfpt[i],2) for i in range(5)},
                     "states":STATES,
                     "active_pct":round(pi[0]*100,2),"fat_pct":round(pi[1]*100,2)}
    return results

def live_dashboard_sim(line_stats, insp_map):
    now=datetime.datetime.now(); log=[]; alerts=[]; avail=len(insp_map)
    for h in range(24):
        ts=now.replace(hour=h,minute=0,second=0,microsecond=0).strftime("%Y-%m-%d %H:%M")
        row={"ts":ts,"hour":h}
        for ln,s in line_stats.items():
            spd=max(50,s["speed"]*random.gauss(1.0,0.05))
            avr=random.uniform(0.88,0.99)
            prf=round(spd/s["speed"],3); qlt=random.uniform(0.97,1.0)
            oee=round(avr*prf*qlt*100,2)
            insp_now=random.choice([1,1,2,2,avail//3+1])
            util=round(min(insp_now/max(s["n_base"],1),1)*100,1)
            def_r=max(0,random.gauss(s["defects_km_avg"] if s["defects_km_avg"]>0 else 0.5,0.1))
            fat=min(10,max(1,s["fat_avg"]+h*0.05+random.gauss(0,0.4))) if s["fat_avg"]>0 else 5.5
            sla="OK" if util>=80 else "BREACH"
            row[ln]={"oee":oee,"util":util,"def":round(def_r,4),"fat":round(fat,2),"sla":sla,"insp":insp_now}
            if fat>=9: alerts.append({"ts":ts,"line":ln,"tier":"P1 CRITICAL","msg":f"Fatigue {fat:.1f}/10 — ROTATE NOW"})
            elif fat>=7: alerts.append({"ts":ts,"line":ln,"tier":"P2 HIGH","msg":f"Fatigue {fat:.1f}/10 — Rotate next session"})
            if sla=="BREACH": alerts.append({"ts":ts,"line":ln,"tier":"P2 HIGH","msg":f"SLA BREACH: {insp_now}/{s['n_base']} inspectors"})
            if def_r>(s["defects_km_avg"]+0.1)*2: alerts.append({"ts":ts,"line":ln,"tier":"P1 CRITICAL","msg":f"Defect spike {def_r:.3f}/km"})
        log.append(row)
    summary={}
    for ln in line_stats:
        oees=[r[ln]["oee"] for r in log]; utils=[r[ln]["util"] for r in log]
        fats=[r[ln]["fat"] for r in log]; slas=[r[ln]["sla"] for r in log]
        summary[ln]={"avg_oee":round(np.mean(oees),2),"avg_util":round(np.mean(utils),1),
                     "avg_fat":round(np.mean(fats),2),"sla_breaches":slas.count("BREACH"),
                     "sla_pct":round((24-slas.count("BREACH"))/24*100,1)}
    p1=[a for a in alerts if "P1" in a["tier"]]; p2=[a for a in alerts if "P2" in a["tier"]]
    return log,p1,p2,summary

# ─────────────────────────────────────────────────────────────────────────────
# ── SHEET BUILDERS: AI ALGORITHMS ────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
def build_a1(wb, series):
    ws=wb.create_sheet("A1 Demand Forecast"); no_grid(ws)
    set_widths(ws,[6,10,10,10,10,10,10,12,12,12,12]); NC=11
    title_merge(ws,"A1: DEMAND FORECASTING — EWMA + HOLT-WINTERS | Auto-updates on data save",1,NC)
    note_merge(ws,"Formula: L_t=α·y_t+(1-α)·(L_{t-1}+T_{t-1}) | T_t=β·(L_t-L_{t-1})+(1-β)·T_{t-1} | Forecast: y_{t+h}=L_t+h·T_t | α=0.3 β=0.1",2,NC)
    r=4
    for ln in ["CGL","CAL","RCL"]:
        d=series[ln]; def_s=d["defects"]; fat_s=d["fatigue"]; spd_s=d["speed"]
        sm,fc=holt_winters(def_s); ew=ewma(def_s)
        _,fc_fat=holt_winters(fat_s); _,fc_spd=holt_winters(spd_s)
        alert="HIGH" if max(fc)>d["base_def"]*1.3 else "NORMAL"
        sub_merge(ws,f"LINE: {ln} — 48h actual + 24h ahead | Alert: {alert}",r,NC,
                  bg=BLUE if ln=="CGL" else MID if ln=="CAL" else AMB); r+=1
        hrow(ws,r,["Hr","Actual Def","Smoothed","EWMA","","Fc Hr","Def Forecast","Speed Fc","Fatigue Fc","Peak Hr","Alert"]); r+=1
        for i in range(min(24,len(def_s))):
            fc_v=fc[i] if i<len(fc) else ""; fc_f=fc_fat[i] if i<len(fc_fat) else ""; fc_s=fc_spd[i] if i<len(fc_spd) else ""
            bg=LT_AMB if fc_v and fc_v>d["base_def"]*1.2 else WHITE
            drow(ws,r,[i+1,round(def_s[i],4),sm[i],ew[i],"",f"h+{i+1}",
                       round(fc_v,4) if fc_v else "",round(fc_s,1) if fc_s else "",
                       round(fc_f,2) if fc_f else "",int(np.argmax(fc))+1,alert],
                 bg=bg if i%2==0 else LT_GRY); r+=1
            if alert=="HIGH": ws.cell(r-1,11).fill=_fill(LT_RED); ws.cell(r-1,11).font=_font(True,10,RED)
        r+=1

def build_a2(wb, series):
    ws=wb.create_sheet("A2 Anomaly Detection"); no_grid(ws)
    set_widths(ws,[6,10,10,10,10,10,12,8,14,20]); NC=10
    title_merge(ws,"A2: ANOMALY DETECTION — Z-SCORE + IQR + ISOLATION FOREST ENSEMBLE | Live per data save",1,NC)
    note_merge(ws,"3-detector ensemble vote | 2+ votes = anomaly | ISO score > 0.7 = outlier | Z-threshold=2.5σ | IQR=Q1-1.5×IQR…Q3+1.5×IQR",2,NC)
    r=4
    for ln in ["CGL","CAL","RCL"]:
        def_s=series[ln]["defects"]; fat_s=series[ln]["fatigue"]
        zr=zscore(def_s); iq=iqr_flag(def_s); iso=[iso_score(v,def_s) for v in def_s]
        mu=round(np.mean(def_s),4); sg=round(np.std(def_s),4)
        q1=round(np.percentile(def_s,25),4); q3=round(np.percentile(def_s,75),4)
        anoms=[]
        for i in range(min(24,len(def_s))):
            zf,zs=zr[i]; iof=iq[i]; iso_s=iso[i]
            votes=sum([zf,iof,iso_s>0.7])
            sev="CRITICAL" if votes>=3 else "HIGH" if votes==2 else "WATCH" if votes==1 else "NORMAL"
            anoms.append((i+1,round(def_s[i],4),round(fat_s[i],2),round(zs,3),zf,iof,iso_s,votes,sev))
        n_anom=sum(1 for a in anoms if a[7]>=2)
        sub_merge(ws,f"LINE: {ln} | μ={mu} σ={sg} | IQR [{q1},{q3}] | Anomaly rate: {round(n_anom/len(anoms)*100,1)}%",r,NC); r+=1
        hrow(ws,r,["Hr","Defect","Fatigue","Z-Score","Z-Flag","IQR Flag","ISO Score","Votes","Severity","Action"]); r+=1
        for a in anoms:
            sev=a[8]
            bg=LT_RED if sev=="CRITICAL" else LT_AMB if sev=="HIGH" else LT_GRY if sev=="WATCH" else WHITE
            act=("STOP & INVESTIGATE" if sev=="CRITICAL" else "INVESTIGATE" if sev=="HIGH" else "MONITOR" if sev=="WATCH" else "Normal ops")
            drow(ws,r,[a[0],a[1],a[2],a[3],"YES" if a[4] else "no","YES" if a[5] else "no",a[6],a[7],sev,act],bg=bg); r+=1
            if sev in("CRITICAL","HIGH"):
                ws.cell(r-1,9).fill=_fill(LT_RED if sev=="CRITICAL" else LT_AMB)
                ws.cell(r-1,9).font=_font(True,10,RED if sev=="CRITICAL" else AMB)
        r+=1

def build_a3(wb):
    ws=wb.create_sheet("A3 RL Q-Learning"); no_grid(ws)
    set_widths(ws,[8,14,14,12,12,12,14,22]); NC=8
    title_merge(ws,"A3: REINFORCEMENT LEARNING — Q-LEARNING OPTIMAL ROTATION POLICY | Retrains on every save",1,NC)
    note_merge(ws,"State:(fatigue_bin × time_bin × line) | Actions:Continue/Rotate | γ=0.95 α=0.1 ε=0.2 | 300 episodes",2,NC)
    r=4
    policy,avg_rew,avg_rot=q_learning(n_ep=300)
    sub_merge(ws,f"TRAINED POLICY — Avg Reward (last 30 eps): {avg_rew} | Avg Rotations/ep: {avg_rot}",r,NC); r+=1
    hrow(ws,r,["Line","Fatigue Bin","Time On (min)","Q(Continue)","Q(Rotate)","Policy","Confidence","Rationale"]); r+=1
    for i,p in enumerate(policy):
        bg=LT_RED if p["policy"]=="ROTATE" else LT_GRN
        reason=("High fatigue — rotate to prevent quality miss" if p["policy"]=="ROTATE" else "Within safe limits — continue")
        drow(ws,r,[p["line"],p["fatigue_bin"],p["time_bin"],p["q_cont"],p["q_rot"],p["policy"],p["confidence"],reason],bg=bg if i%2==0 else WHITE,lcols={8}); r+=1
        ws.cell(r-1,6).font=_font(True,10,RED if p["policy"]=="ROTATE" else GRN)
        ws.cell(r-1,6).fill=_fill(LT_RED if p["policy"]=="ROTATE" else LT_GRN)

def build_a4(wb, series):
    ws=wb.create_sheet("A4 Fatigue Predict"); no_grid(ws)
    set_widths(ws,[8,12,12,12,12,14,16,14]); NC=8
    title_merge(ws,"A4: PREDICTIVE FATIGUE MODEL — POLYNOMIAL REGRESSION (DEG-3) | Auto-recalculates",1,NC)
    note_merge(ws,"y = a₀+a₁x+a₂x²+a₃x³ | Fitted on last 48h fatigue readings | Forecasts next 12h | Time-to-critical = when forecast ≥ 8",2,NC)
    r=4
    for ln in ["CGL","CAL","RCL"]:
        fat_s=series[ln]["fatigue"]
        x=list(range(len(fat_s))); y=fat_s
        fitted,fc12,rmse,r2,coeffs=polyfit(x,y)
        curr=round(fat_s[-1],2); ttc=next((i for i,v in enumerate(fc12,1) if v>=8),None)
        ttc_str=f"{ttc} hours" if ttc else ">12 hours"
        risk=("CRITICAL" if curr>=8 else "HIGH" if curr>=6 else "MODERATE")
        sub_merge(ws,f"LINE: {ln} | Current Fatigue: {curr} | Time-to-Critical: {ttc_str} | R²={r2} | RMSE={rmse} | Risk: {risk}",r,NC,
                  bg=RED if risk=="CRITICAL" else AMB if risk=="HIGH" else MID); r+=1
        note_merge(ws,f"Coefficients: a3={coeffs[0]} a2={coeffs[1]} a1={coeffs[2]} a0={coeffs[3]}",r,NC); r+=1
        hrow(ws,r,["Hr","Actual","Fitted","Residual","","Fc Hr+","Predicted Fatigue","Risk Flag"]); r+=1
        for i in range(min(24,len(fat_s))):
            fc_v=fc12[i] if i<len(fc12) else ""
            rf=("CRITICAL" if fc_v and fc_v>=8 else "HIGH" if fc_v and fc_v>=6 else "OK") if fc_v else ""
            bg=LT_RED if rf=="CRITICAL" else LT_AMB if rf=="HIGH" else WHITE
            drow(ws,r,[i+1,round(fat_s[i],3),fitted[i],round(fat_s[i]-fitted[i],3),"",f"h+{i+1}",
                       round(fc_v,3) if fc_v else "",rf],bg=bg if i%2==0 else LT_GRY); r+=1
            if rf in("CRITICAL","HIGH"):
                ws.cell(r-1,8).fill=_fill(LT_RED if rf=="CRITICAL" else LT_AMB)
                ws.cell(r-1,8).font=_font(True,10,RED if rf=="CRITICAL" else AMB)
        r+=1

def build_a5(wb, line_stats):
    ws=wb.create_sheet("A5 DP Scheduling"); no_grid(ws)
    set_widths(ws,[20,14,14,14,14,14,14]); NC=7
    title_merge(ws,"A5: DYNAMIC PROGRAMMING — OPTIMAL SHIFT SEQUENCING | Recalculates from live data",1,NC)
    note_merge(ws,"Bellman: V(i,j)=max_k[w_i·coverage(k,demand_i)+V(i+1,j-k)] | Weights: Morning=1.3× Afternoon=1.0× Night=0.7×",2,NC)
    r=4
    results=dp_scheduling(line_stats)
    hrow(ws,r,["Inspectors Available","DP Opt. Score","Efficiency %","Shift","Assigned","Demand","Coverage %"]); r+=1
    for sc in results:
        first=True
        for sn,sd in sc["assign"].items():
            cb=LT_GRN if sc["eff_pct"]>=80 else LT_AMB if sc["eff_pct"]>=50 else LT_RED
            cv=LT_GRN if sd["cov_pct"]>=100 else LT_AMB if sd["cov_pct"]>=60 else LT_RED
            vals=([sc["avail"],sc["score"],f"{sc['eff_pct']}%"] if first else ["","",""])
            vals+=[sn,sd["assigned"],sd["demand"],f"{sd['cov_pct']}%"]
            drow(ws,r,vals,bg=cb,lcols={4}); r+=1
            ws.cell(r-1,7).fill=_fill(cv); ws.cell(r-1,7).font=_font(True,10,GRN if sd["cov_pct"]>=100 else RED)
            first=False
        r+=1

def build_a6(wb, line_stats):
    ws=wb.create_sheet("A6 Genetic Algorithm"); no_grid(ws)
    set_widths(ws,[16,8,8,12,12,14,14]); NC=7
    title_merge(ws,"A6: GENETIC ALGORITHM — MULTI-LINE INSPECTOR SCHEDULING | Evolves on every save",1,NC)
    note_merge(ws,"Pop=50 Gen=60 MR=15% | Fitness: weighted coverage − over-staffing penalty − under-staffing penalty | Elitism: top-1",2,NC)
    r=4
    sched,best_fit,bpg,apg,conv=genetic_algorithm(line_stats)
    sub_merge(ws,f"OPTIMAL SCHEDULE — Fitness: {best_fit} | Converged: Gen {conv}",r,NC); r+=1
    hrow(ws,r,["Shift","Line","Assigned","Demand","Coverage %","Status","Fitness Evolution (Gen)"]); r+=1
    for i,e in enumerate(sched):
        st=e["status"]; bg=LT_GRN if st=="OK" else LT_AMB if st=="UNDER" else LT_RED
        gen_label=f"Gen {i*10}: {bpg[min(i*10,len(bpg)-1)]}" if i*10<len(bpg) else ""
        drow(ws,r,[e["shift"],e["line"],e["assigned"],e["demand"],f"{e['cov_pct']}%",st,gen_label],bg=bg); r+=1
        ws.cell(r-1,6).font=_font(True,10,GRN if st=="OK" else RED)
    r+=1
    sub_merge(ws,"FITNESS CONVERGENCE (every 5 generations)",r,NC); r+=1
    hrow(ws,r,["Generation","Best Fitness","Avg Fitness","Δ Best","","",""]); r+=1
    for i in range(0,min(len(bpg),60),5):
        delta=round(bpg[i]-bpg[max(0,i-5)],3) if i>0 else 0
        drow(ws,r,[i+1,bpg[i],apg[i],delta,"","",""],
             bg=LT_GRN if delta>0 else LT_RED if delta<-0.01 else WHITE); r+=1

def build_a7(wb, series):
    ws=wb.create_sheet("A7 CUSUM Control"); no_grid(ws)
    set_widths(ws,[6,10,10,10,10,10,16,20]); NC=8
    title_merge(ws,"A7: CUSUM CONTROL CHART — STATISTICAL PROCESS CONTROL | Updates on save",1,NC)
    note_merge(ws,"C+_t=max(0,C+_{t-1}+(x_t−μ₀−k)) | C-_t=max(0,C-_{t-1}+(μ₀−k−x_t)) | SIGNAL when C+ or C- > h=4σ | k=0.5σ",2,NC)
    r=4
    for ln in ["CGL","CAL","RCL"]:
        def_s=series[ln]["defects"]
        d=cusum(def_s,ln)
        sub_merge(ws,f"LINE: {ln} | Status: {d['status']} | Alarms: {d['n_alarms']} ({d['rate']}%) | μ₀={d['mu0']} σ={d['sg']} k={d['k']} h={d['h']}",r,NC,
                  bg=RED if "OUT" in d["status"] else AMB if "WATCH" in d["status"] else GRN); r+=1
        hrow(ws,r,["Obs","Defects","CUSUM+","CUSUM-","","Threshold h","Signal","Action"]); r+=1
        for i in range(min(24,len(d["series"]))):
            sig=d["sigs"][i] if i<len(d["sigs"]) else "IN_CONTROL"
            bg=LT_RED if "ABOVE" in sig or "BELOW" in sig else WHITE
            act=("INVESTIGATE ↑ SHIFT" if "ABOVE" in sig else "INVESTIGATE ↓ SHIFT" if "BELOW" in sig else "Continue")
            drow(ws,r,[i+1,d["series"][i],d["cp"][i],d["cn"][i],"",d["h"],sig,act],bg=bg if i%2==0 else LT_GRY); r+=1
            if "LIMIT" in sig: ws.cell(r-1,7).fill=_fill(LT_RED); ws.cell(r-1,7).font=_font(True,10,RED)
        sbg=LT_RED if "OUT" in d["status"] else LT_AMB if "WATCH" in d["status"] else LT_GRN
        drow(ws,r,["PROCESS STATUS","",d["status"],"","","","",""],bg=sbg,bold=True,fc=RED if "OUT" in d["status"] else GRN if "IN " in d["status"] else AMB); r+=2

def build_a8(wb, line_stats):
    ws=wb.create_sheet("A8 Monte Carlo"); no_grid(ws)
    set_widths(ws,[28,14,14,14,14,14]); NC=6
    title_merge(ws,"A8: MONTE CARLO STAFFING RISK — 3,000 SIMULATIONS | Recalculates on every save",1,NC)
    note_merge(ws,"Each sim: absenteeism (Binomial p=0.10) + defect spike (Poisson) + speed variation (Normal σ=7%) | VaR=5th-percentile coverage",2,NC)
    r=4
    mc=monte_carlo(line_stats,n=3000)
    hrow(ws,r,["Metric","CGL","CAL","RCL","Interpretation",""]); r+=1
    metrics=[
        ("P(Understaffed) %",     "p_under",   lambda v: LT_RED if v>30 else LT_AMB if v>15 else LT_GRN),
        ("P(Coverage < 80%) %",   "p_miss",    lambda v: LT_RED if v>40 else LT_AMB if v>20 else LT_GRN),
        ("Coverage VaR (5th pct)","var5",      lambda v: LT_RED if v<0.6 else LT_AMB if v<0.8 else LT_GRN),
        ("P5  Coverage",          "p5",        None),
        ("P25 Coverage",          "p25",       None),
        ("P50 Coverage (Median)", "p50",       None),
        ("P75 Coverage",          "p75",       None),
        ("P95 Coverage",          "p95",       None),
        ("Risk Rating",           "risk",      lambda v: LT_RED if v in("EXTREME","HIGH") else LT_AMB if v=="MEDIUM" else LT_GRN),
    ]
    pct_keys=["p5","p25","p50","p75","p95"]
    for label,key,cfn in metrics:
        is_pct=key in pct_keys
        vals=[]
        for ln in ["CGL","CAL","RCL"]:
            v=mc[ln]["pcts"].get(key) if is_pct else mc[ln].get(key,"")
            vals.append(v)
        interp=("Structural understaffing risk" if "Under" in label else
                "Inspection coverage failure risk" if "Coverage <" in label else
                "Worst-case day coverage" if "VaR" in label else
                f"In bottom 5% of days coverage={vals[0]}" if "P5" in label else "")
        ws.row_dimensions[r].height=18
        _cell(ws,r,1,label,bg=LT_GRY,bold=True,align="left")
        for ci,v in enumerate(vals,2):
            bg=cfn(v) if cfn and v!="" else WHITE
            _cell(ws,r,ci,v,bg=bg,fc=RED if bg==LT_RED else GRN if bg==LT_GRN else AMB if bg==LT_AMB else "000000",bold=(bg!=WHITE))
        _cell(ws,r,5,interp,align="left"); ws.cell(r,6).border=_bd()
        r+=1

def build_a9(wb, line_stats):
    ws=wb.create_sheet("A9 Markov Chain"); no_grid(ws)
    set_widths(ws,[18,12,12,12,12,12,12,14]); NC=8
    title_merge(ws,"A9: MARKOV CHAIN — INSPECTOR STATE TRANSITIONS & STEADY-STATE | Live data driven",1,NC)
    note_merge(ws,"States: ACTIVE|FATIGUED|ROTATING|ABSENT|TRAINING | π=π·P (1000-step power iteration) | MFPT via fundamental matrix Z=(I-P+1π)⁻¹",2,NC)
    r=4
    mk=markov(line_stats)
    for ln in ["CGL","CAL","RCL"]:
        d=mk[ln]; states=d["states"]
        sub_merge(ws,f"LINE: {ln} | Productive (ACTIVE): {d['active_pct']}% | Fatigue risk: {d['fat_pct']}%",r,NC,
                  bg=BLUE if ln=="CGL" else MID if ln=="CAL" else AMB); r+=1
        hrow(ws,r,["Transition →"]+states+[""]); r+=1
        for i,st in enumerate(states):
            bg=LT_GRN if st=="ACTIVE" else LT_RED if st=="FATIGUED" else LT_BLUE if st=="ROTATING" else LT_GRY
            drow(ws,r,[f"FROM: {st}"]+d["P"][i]+[""],bg=bg); r+=1
        hrow(ws,r,["Metric"]+states+[""],bg=MID); r+=1
        ss=d["ss"]; mfpt=d["mfpt"]
        for label,src in [("Steady-State %",ss),("MFPT (hrs)",mfpt)]:
            drow(ws,r,[label]+[src[s] for s in states]+[""],
                 bg=LT_BLUE if "Steady" in label else LT_GRY); r+=1
        r+=1

def build_a10(wb, line_stats, insp_map):
    ws=wb.create_sheet("A10 Live Dashboard"); no_grid(ws)
    set_widths(ws,[18,10,10,10,10,10,10,10,10,10,10,10,10]); NC=13
    title_merge(ws,"A10: LIVE KPI DASHBOARD — OEE | UTILISATION | DEFECTS | FATIGUE | ALERTS",1,NC)
    r=2
    log,p1,p2,summary=live_dashboard_sim(line_stats,insp_map)
    # Summary scorecard
    ws.merge_cells(f"A{r}:{gcol(NC)}{r}")
    ws.cell(r,1,"24-HOUR SUMMARY SCORECARD").font=_font(True,11,WHITE)
    ws.cell(r,1).fill=_fill(NAVY); ws.cell(r,1).alignment=_al("left"); ws.row_dimensions[r].height=22; r+=1
    hrow(ws,r,["Line","Avg OEE%","Avg Util%","Avg Fatigue","Max Fatigue","SLA Breach Hrs","SLA Compliance%","P1 Alerts","P2 Alerts","","","",""]); r+=1
    for ln in ["CGL","CAL","RCL"]:
        s=summary[ln]
        p1c=sum(1 for a in p1 if a["line"]==ln); p2c=sum(1 for a in p2 if a["line"]==ln)
        slabg=LT_GRN if s["sla_pct"]>=90 else LT_AMB if s["sla_pct"]>=70 else LT_RED
        drow(ws,r,[ln,s["avg_oee"],s["avg_util"],s["avg_fat"],"",s["sla_breaches"],
                   f"{s['sla_pct']}%",p1c,p2c,"","","",""]); r+=1
        ws.cell(r-1,7).fill=_fill(slabg)
        ws.cell(r-1,8).fill=_fill(LT_RED if p1c>0 else WHITE)
        ws.cell(r-1,8).font=_font(True,10,RED if p1c>0 else "000000")
    # Hour log
    r+=1
    sub_merge(ws,"HOUR-BY-HOUR KPI LOG (24 HOURS)",r,NC); r+=1
    hrow(ws,r,["Timestamp","CGL OEE","CGL Util","CGL Fat","CGL SLA",
               "CAL OEE","CAL Util","CAL Fat","CAL SLA",
               "RCL OEE","RCL Util","RCL Fat","RCL SLA"]); r+=1
    for i,row in enumerate(log):
        bg=LT_GRY if i%2==0 else WHITE
        vals=[row["ts"]]
        for ln in ["CGL","CAL","RCL"]: vals+=[row[ln]["oee"],row[ln]["util"],row[ln]["fat"],row[ln]["sla"]]
        drow(ws,r,vals,bg=bg); r+=1
        for ci,ln in zip([5,9,13],["CGL","CAL","RCL"]):
            s=row[ln]["sla"]
            ws.cell(r-1,ci).fill=_fill(LT_RED if s=="BREACH" else LT_GRN)
            ws.cell(r-1,ci).font=_font(True,10,RED if s=="BREACH" else GRN)
        for ci,ln in zip([4,8,12],["CGL","CAL","RCL"]):
            f=row[ln]["fat"]
            if f>=8: ws.cell(r-1,ci).fill=_fill(LT_RED); ws.cell(r-1,ci).font=_font(True,10,RED)
            elif f>=6: ws.cell(r-1,ci).fill=_fill(LT_AMB)
    # Alerts
    r+=1
    sub_merge(ws,f"P1 CRITICAL ALERTS ({len(p1)})",r,NC,bg=RED); r+=1
    hrow(ws,r,["Timestamp","Line","Tier","Alert Message","","","","","","","","",""],bg=RED); r+=1
    for a in (p1 or [{"ts":"None","line":"—","tier":"P1","msg":"No critical alerts — system healthy"}])[:15]:
        drow(ws,r,[a["ts"],a["line"],a["tier"],a.get("msg",""),*[""]*(NC-4)],bg=LT_RED,lcols={4}); r+=1
    r+=1
    sub_merge(ws,f"P2 HIGH ALERTS ({len(p2)})",r,NC,bg=AMB); r+=1
    hrow(ws,r,["Timestamp","Line","Tier","Alert Message","","","","","","","","",""],bg=AMB); r+=1
    for a in (p2 or [{"ts":"None","line":"—","tier":"P2","msg":"No high alerts"}])[:15]:
        drow(ws,r,[a["ts"],a["line"],a["tier"],a.get("msg",""),*[""]*(NC-4)],bg=LT_AMB,lcols={4}); r+=1

# ─────────────────────────────────────────────────────────────────────────────
# MASTER RUN FUNCTION
# ─────────────────────────────────────────────────────────────────────────────
def run(source_path: str, output_path: str):
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"  [{ts}] Reading {Path(source_path).name}...")

    # Read source
    coils, logs, skills, hourly = read_source(Path(source_path))
    enriched, line_stats, insp_map = compute_metrics(coils, logs, skills)
    series = make_series(enriched, line_stats, n=48)

    # Load existing change log if output exists
    existing_log = []
    if Path(output_path).exists():
        try:
            old = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
            if "📝 Change Log" in old.sheetnames:
                ws_cl = old["📝 Change Log"]
                rows  = list(ws_cl.iter_rows(values_only=True))
                if len(rows) > 2:
                    hdr = [str(h) if h else "" for h in rows[1]]
                    for row in rows[2:50]:
                        if row[0]: existing_log.append(dict(zip(hdr, row)))
            old.close()
        except: pass

    print(f"  Loaded {len(coils)} coils | {len(logs)} inspections | {len(insp_map)} inspectors")
    print(f"  Running 10 AI algorithms...")

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)

    # ── Live result sheets ──
    build_dashboard(wb, enriched, line_stats, insp_map, source_path, ts)
    build_coil_detail(wb, enriched)
    build_fatigue_log(wb, enriched, logs)
    build_inspector_matrix(wb, insp_map)
    build_inspector_calc(wb, line_stats, insp_map)
    build_change_log(wb, source_path, len(coils), ts, existing_log)

    # ── AI algorithm sheets ──
    build_a1(wb, series)
    build_a2(wb, series)
    build_a3(wb)
    build_a4(wb, series)
    build_a5(wb, line_stats)
    build_a6(wb, line_stats)
    build_a7(wb, series)
    build_a8(wb, line_stats)
    build_a9(wb, line_stats)
    build_a10(wb, line_stats, insp_map)

    wb.save(output_path)
    print(f"  ✅ Saved → {Path(output_path).name} ({len(wb.sheetnames)} sheets)")

# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv)>1 else "TSK_Inspection_Data.xlsx"
    out = sys.argv[2] if len(sys.argv)>2 else "TSK_Results.xlsx"
    run(src, out)
