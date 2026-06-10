# excel_styles.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import numpy as np

# ========== STYLE CONSTANTS ==========
NAVY = "0D2137"
BLUE = "1F4E79"
MID = "2E75B6"
LT_BLUE = "DEEAF1"
RED = "C00000"
LT_RED = "FFE7E7"
AMB = "E36C09"
LT_AMB = "FFF2CC"
GRN = "375623"
LT_GRN = "EBF3E8"
GREY = "595959"
WHITE = "FFFFFF"
ACC = "00B0D7"
LT_GRY = "F5F5F5"
OFF_W = "F7F9FC"
DARK = "0A1628"

def _s(c="CCCCCC", t="thin"):
    return Side(style=t, color=c)

def _fill(c):
    return PatternFill("solid", fgColor=c)

def _font(b=False, sz=10, c="000000"):
    return Font(bold=b, size=sz, color=c, name="Calibri")

def _al(h="center", v="center", w=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=w)

def _bd(c="CCCCCC"):
    return Border(left=_s(c), right=_s(c), top=_s(c), bottom=_s(c))

def _cell(ws, r, c, val, bg=WHITE, fc="000000", bold=False, sz=10, align="center", border=True, fmt=None):
    cl = ws.cell(row=r, column=c, value=val)
    cl.font = _font(bold, sz, fc)
    cl.fill = _fill(bg)
    cl.alignment = _al(align)
    if border:
        cl.border = _bd()
    if fmt:
        cl.number_format = fmt
    return cl

def hrow(ws, row, vals, bg=NAVY, fc=WHITE, bold=True, h=26, start=1):
    ws.row_dimensions[row].height = h
    for i, v in enumerate(vals, start):
        _cell(ws, row, i, v, bg=bg, fc=fc, bold=bold, align="center")

def drow(ws, row, vals, bg=WHITE, fc="000000", bold=False, h=18, lcols=None, start=1):
    ws.row_dimensions[row].height = h
    for i, v in enumerate(vals, start):
        al = "left" if lcols and i in lcols else "center"
        _cell(ws, row, i, v, bg=bg, fc=fc, bold=bold, align=al)

def title_merge(ws, text, row, end_col, bg=NAVY, fc=WHITE, sz=14, h=36):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cl = ws.cell(row=row, column=1, value=text)
    cl.font = _font(True, sz, fc)
    cl.fill = _fill(bg)
    cl.alignment = _al("center")
    ws.row_dimensions[row].height = h

def sub_merge(ws, text, row, end_col, bg=BLUE, fc=WHITE, sz=10, h=20):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cl = ws.cell(row=row, column=1, value=text)
    cl.font = _font(True, sz, fc)
    cl.fill = _fill(bg)
    cl.alignment = _al("left")
    ws.row_dimensions[row].height = h

def note_merge(ws, text, row, end_col, bg=OFF_W, h=20):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cl = ws.cell(row=row, column=1, value=text)
    cl.font = _font(False, 9, GREY)
    cl.fill = _fill(bg)
    cl.alignment = _al("left")
    ws.row_dimensions[row].height = h

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def no_grid(ws):
    ws.sheet_view.showGridLines = False

# -------------------- Sheet Builders --------------------
def build_dashboard(wb, coils, inspections, inspectors, line_stats):
    ws = wb.create_sheet("🏠 Dashboard")
    set_widths(ws, [18,14,14,14,14,14,14,14,14,14,14,14])
    NC = 12
    title_merge(ws, "TSK COIL INSPECTION — LIVE RESULTS DASHBOARD", 1, NC, sz=16)
    ws.cell(2,1).value = f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Source: Database  |  Auto-generated"
    ws.cell(2,1).font = _font(False, 9, ACC)
    ws.cell(2,1).fill = _fill(NAVY)
    ws.cell(2,1).alignment = _al("center")
    ws.row_dimensions[2].height = 18

    total_coils = len(coils)
    avail_insp = len(inspectors)
    base_req = sum(v["n_base"] for v in line_stats.values())
    peak_req = sum(v["n_peak"] for v in line_stats.values())
    gap = max(0, base_req - avail_insp)
    all_fats = [i.fatigue_score_post for i in inspections if i.fatigue_score_post]
    avg_fat = round(np.mean(all_fats), 2) if all_fats else 0

    kpis = [
        (str(total_coils),   "Total Coils\nInspected",     BLUE,    LT_BLUE,  "000000"),
        (str(avail_insp),    "Inspectors\nAvailable",      GRN if avail_insp>=base_req else RED, LT_GRN if avail_insp>=base_req else LT_RED, "000000"),
        (str(base_req),      "Base\nRequired",             MID,     LT_BLUE,  "000000"),
        (str(peak_req),      "Peak\nRequired",             AMB,     LT_AMB,   "000000"),
        (f"−{gap}" if gap>0 else "0", "Staffing\nGap",   RED if gap>0 else GRN, LT_RED if gap>0 else LT_GRN, RED if gap>0 else GRN),
        (f"{avg_fat}",       "Avg Fatigue\n(target ≤6)",  RED if avg_fat>7 else AMB if avg_fat>5 else GRN, LT_RED if avg_fat>7 else LT_AMB, "000000"),
    ]
    r=4
    for col_s, (val, lbl, bg, lbg, fc) in enumerate(kpis, 1):
        ws.merge_cells(start_row=r, start_column=col_s*2-1, end_row=r, end_column=col_s*2)
        cl = ws.cell(r, col_s*2-1, val)
        cl.font = _font(True,28,bg)
        cl.fill = _fill(lbg)
        cl.alignment = _al("center")
        cl.border = _bd(bg)
        ws.merge_cells(start_row=r+1, start_column=col_s*2-1, end_row=r+1, end_column=col_s*2)
        cl2 = ws.cell(r+1, col_s*2-1, lbl)
        cl2.font = _font(False,9,GREY)
        cl2.fill = _fill(lbg)
        cl2.alignment = _al("center")
        cl2.border = _bd(bg)
    ws.row_dimensions[r].height = 40
    ws.row_dimensions[r+1].height = 26

    # Per‑line table
    r=7
    sub_merge(ws, "PER-LINE PERFORMANCE METRICS", r, NC); r+=1
    hrow(ws, r, ["Line","Speed (m/min)","Coils","Avg W_l (s/m)","Min W_l","Max W_l",
                 "Avg Def/km","CV","Base Insp.","Peak Insp.","Fatigue Avg","Status"]); r+=1
    for i, (ln, s) in enumerate(line_stats.items()):
        ok = s["n_base"] <= avail_insp
        status = "OK" if ok else "GAP"
        sbg = LT_GRN if ok else LT_RED
        sfc = GRN if ok else RED
        bgs = [LT_BLUE if i%2==0 else WHITE]*12
        drow(ws, r, [ln, s["speed"], s["n_coils"], s["wl_avg"], s.get("wl_min",0), s.get("wl_max",0),
                     s["defects_km_avg"], s["cv"], s["n_base"], s["n_peak"], s["fat_avg"], status],
             bg=bgs[0]); r+=1
        ws.cell(r-1,12).fill = _fill(sbg)
        ws.cell(r-1,12).font = _font(True,10,sfc)

    # Workforce gap table
    r+=1
    sub_merge(ws, "WORKFORCE GAP SUMMARY", r, NC); r+=1
    hrow(ws, r, ["Line","Available","Base Req.","Peak Req.","Gap (Base)","Gap (Peak)",
                 "% Shortfall","Priority","Action Required","Rotation (min)","Fatigue Risk",""]); r+=1
    rot_map = {"CGL":45,"CAL":60,"RCL":30}
    for i, (ln, s) in enumerate(line_stats.items()):
        av2 = avail_insp
        gb = max(0, s["n_base"] - av2)
        gp = max(0, s["n_peak"] - av2)
        pct = round(gb/s["n_base"]*100) if s["n_base"]>0 else 0
        prio = "CRITICAL" if gb>2 else "HIGH" if gb>0 else "OK"
        action = "RECRUIT IMMEDIATELY" if gb>1 else "MONITOR" if gb>0 else "Maintain"
        fat_str = f"{s['fat_avg']}/10 {'CRITICAL' if s['fat_avg']>=8 else 'HIGH' if s['fat_avg']>=6 else 'OK'}"
        pbg = LT_RED if "CRITICAL" in prio else LT_AMB if "HIGH" in prio else LT_GRN
        pfc = RED if "CRITICAL" in prio else AMB if "HIGH" in prio else GRN
        drow(ws, r, [ln, av2, s["n_base"], s["n_peak"], f"−{gb}" if gb>0 else "0",
                     f"−{gp}" if gp>0 else "0", f"{pct}%", prio, action,
                     rot_map.get(ln,45), fat_str, ""],
             bg=LT_BLUE if i%2==0 else WHITE); r+=1
        ws.cell(r-1,8).fill = _fill(pbg)
        ws.cell(r-1,8).font = _font(True,10,pfc)
    drow(ws, r, ["TOTAL", avail_insp, base_req, peak_req,
                 f"−{gap}" if gap>0 else "0",
                 f"−{max(0,peak_req-avail_insp)}" if peak_req>avail_insp else "0",
                 "","","","","",""],
         bg=NAVY, fc=WHITE, bold=True)

def build_coil_detail(wb, coils, inspections):
    ws = wb.create_sheet("📋 Coil Detail")
    set_widths(ws, [12,12,12,12,12,12,12,12,14,12])
    title_merge(ws, "COIL-BY-COIL INSPECTION DETAIL", 1, 10)
    hrow(ws, 2, ["Coil ID","Line","Length (m)","Speed (m/s)","Defect Count","Defects/km","W_l (s/m)","Inspector","Duration (min)","Fatigue Score"])
    
    # Build a mapping from coil id to inspection (assumes one inspection per coil)
    insp_by_coil = {insp.coil_id: insp for insp in inspections}
    
    r = 3
    for i, c in enumerate(coils):
        # Defects/km
        defects_km = (c.defect_count / (c.length_m / 1000)) if c.length_m else 0
        defects_km_rounded = round(defects_km, 4)
        
        insp = insp_by_coil.get(c.id)
        if insp:
            inspector = insp.inspector_id
            duration = (insp.inspection_end - insp.inspection_start).total_seconds() / 60.0 if insp.inspection_start and insp.inspection_end else 0
            duration_rounded = round(duration, 1)
            fatigue = insp.fatigue_score_post or 0
            wl = (duration_rounded * 60) / c.length_m if c.length_m else 0
            wl_rounded = round(wl, 4)
        else:
            inspector = ""
            duration_rounded = 0
            fatigue = 0
            wl_rounded = 0
        
        drow(ws, r, [c.coil_id, c.line, c.length_m, c.speed_mps, c.defect_count,
                     defects_km_rounded, wl_rounded, inspector, duration_rounded, fatigue],
             bg=LT_GRY if i%2==0 else WHITE)
        r += 1

def build_fatigue_log(wb, inspections, coils):
    ws = wb.create_sheet("😴 Fatigue Log")
    set_widths(ws, [14,12,14,10,12,22,10])
    title_merge(ws, "FATIGUE SCORE LOG — ALL INSPECTIONS", 1, 7)
    hrow(ws, 2, ["Inspection ID","Coil ID","Inspector ID","Score","Band","Action Required","Risk"])
    r=3
    for i, insp in enumerate(inspections):
        coil = next((c for c in coils if c.id == insp.coil_id), None)
        coil_id = coil.coil_id if coil else "Unknown"
        fat = insp.fatigue_score_post or 0
        band = "CRITICAL" if fat>=8 else "HIGH" if fat>=6 else "MODERATE" if fat>=4 else "LOW"
        action = "Immediate rotation" if fat>=9 else "Rotate next shift" if fat>=7 else "Monitor" if fat>=5 else "Normal"
        bg = LT_RED if fat>=8 else LT_AMB if fat>=6 else LT_GRN if fat>0 else WHITE
        fc = RED if fat>=8 else AMB if fat>=6 else GRN if fat>0 else "000000"
        drow(ws, r, [insp.id, coil_id, insp.inspector_id, fat, band, action, "●"],
             bg=LT_GRY if i%2==0 else WHITE)
        for col in [5,7]:
            ws.cell(r, col).fill = _fill(bg)
            ws.cell(r, col).font = _font(True,10,fc)
        r+=1

def build_inspector_matrix(wb, inspectors):
    ws = wb.create_sheet("👷 Inspector Matrix")
    set_widths(ws, [14,18,10,10,10,14,12,12])
    title_merge(ws, "INSPECTOR CERTIFICATION MATRIX", 1, 8)
    hrow(ws, 2, ["Inspector ID","Name","CGL Cert.","CAL Cert.","RCL Cert.","Lines Certified","Shift Pref.","Risk Level"])
    r=3
    for i, ins in enumerate(inspectors):
        certs = ins.certified_lines or []
        missing = [ln for ln in ["CGL","CAL","RCL"] if ln not in certs]
        risk = "MEDIUM" if len(missing)==1 else "HIGH" if len(missing)==2 else "LOW"
        rbg = LT_AMB if risk=="MEDIUM" else LT_RED if risk=="HIGH" else LT_GRN
        rfc = AMB if risk=="MEDIUM" else RED if risk=="HIGH" else GRN
        drow(ws, r, [ins.inspector_id, ins.name,
                     "YES" if "CGL" in certs else "NO",
                     "YES" if "CAL" in certs else "NO",
                     "YES" if "RCL" in certs else "NO",
                     len(certs), ins.shift_preference, risk],
             bg=LT_GRY if i%2==0 else WHITE, lcols={2})
        for col in [3,4,5]:
            v = ws.cell(r, col).value
            ws.cell(r, col).fill = _fill(LT_GRN if v=="YES" else LT_RED)
            ws.cell(r, col).font = _font(True,10,GRN if v=="YES" else RED)
        ws.cell(r, 8).fill = _fill(rbg)
        ws.cell(r, 8).font = _font(True,10,rfc)
        r+=1

def build_inspector_calc(wb, line_stats):
    ws = wb.create_sheet("🔢 Inspector Calc")
    set_widths(ws, [16,14,14,14,14,14,14,14,20])
    title_merge(ws, "INSPECTOR HEADCOUNT CALCULATOR", 1, 9)
    hrow(ws, 2, ["Line","Speed (m/min)","Avg W_l (s/m)","Defects/km","Def. Burden","Raw Inspectors","Base (Rounded)","Peak (×1.3)","Inspectors Needed"])
    r=3
    for i, (ln, s) in enumerate(line_stats.items()):
        raw = round((s["speed"] * s["wl_avg"]) / 60, 3) if s["wl_avg"]>0 else 2.2
        burden = round(s["defects_km_avg"] / 10, 4)
        needed = f"{s['n_base']} base / {s['n_peak']} peak"
        drow(ws, r, [ln, s["speed"], s["wl_avg"], s["defects_km_avg"], burden, raw, s["n_base"], s["n_peak"], needed],
             bg=LT_BLUE if i%2==0 else WHITE)
        ws.cell(r, 9).fill = _fill(LT_RED if s["n_base"]>3 else LT_GRN)  # example threshold
        r+=1
    base_t = sum(v["n_base"] for v in line_stats.values())
    peak_t = sum(v["n_peak"] for v in line_stats.values())
    drow(ws, r, ["TOTAL", "—", "—", "—", "—", "—", base_t, peak_t, f"Gap: {max(0,base_t-3)} to recruit"],
         bg=NAVY, fc=WHITE, bold=True)

def build_change_log(wb, total_coils, source_file="TSK_Inspection_Data.xlsx"):
    ws = wb.create_sheet("📝 Change Log")
    set_widths(ws, [22,32,28,14,12])
    title_merge(ws, "AUTO-UPDATE CHANGE LOG", 1, 5)
    hrow(ws, 2, ["Timestamp","Event","Source File","Total Coils","Status"])
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    drow(ws, 3, [now, "Auto-recalculation triggered", source_file, total_coils, "SUCCESS"], bg=LT_GRN)