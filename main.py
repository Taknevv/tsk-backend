import os
import tempfile
import math
import random
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from typing import List, Optional

from fastapi import FastAPI, Depends, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.responses import FileResponse
from jose import jwt, JWTError
from sqlalchemy.orm import Session
from dotenv import load_dotenv
from openpyxl import Workbook
from pydantic import BaseModel
from fpdf import FPDF

from database import get_db, engine
from excel_styles import (
    build_dashboard, build_coil_detail, build_fatigue_log,
    build_inspector_matrix, build_inspector_calc, build_change_log,
    no_grid
)
from models import Base, User, Coil, Inspection, Inspector, DefectType, InspectionDefect, Schedule, AuditLog
from schemas import UserCreate, UserOut, Token, CoilCreate, CoilOut, InspectionCreate, InspectionOut
from auth import get_password_hash, verify_password, create_access_token, SECRET_KEY, ALGORITHM

# Firebase Admin
import firebase_admin
from firebase_admin import credentials, messaging

load_dotenv()

# Firebase initialization
cred_path = os.getenv("FIREBASE_CREDENTIALS")
if cred_path and os.path.exists(cred_path):
    cred = credentials.Certificate(cred_path)
    firebase_admin.initialize_app(cred)
    print("Firebase Admin initialized")
else:
    print("Firebase credentials not found – push notifications disabled")

Base.metadata.create_all(bind=engine)

app = FastAPI(title="TSK Coil Backend")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="login")

def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if email is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    user = db.query(User).filter(User.email == email).first()
    if user is None:
        raise credentials_exception
    return user

# ---------- Auth ----------
@app.post("/login", response_model=Token)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == form_data.username).first()
    if not user or not verify_password(form_data.password, user.hashed_password):
        raise HTTPException(status_code=400, detail="Incorrect email or password")
    access_token = create_access_token(data={"sub": user.email})
    return {"access_token": access_token, "token_type": "bearer"}

@app.post("/register", response_model=UserOut)
def register(user: UserCreate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    if current_user.role != 'ua':
        raise HTTPException(status_code=403, detail="Only Ultimate Admin can register users")
    hashed = get_password_hash(user.password)
    db_user = User(
        email=user.email,
        hashed_password=hashed,
        role=user.role,
        line=user.line,
        inspector_id=user.inspector_id,
        name=user.name
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user

@app.get("/me", response_model=UserOut)
def get_me(current_user: User = Depends(get_current_user)):
    return current_user

class FCMToken(BaseModel):
    token: str

@app.post("/user/fcm_token")
def register_fcm_token(data: FCMToken, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    current_user.fcm_token = data.token
    db.commit()
    return {"message": "Token saved"}

@app.get("/defect_types")
def get_defect_types(line: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(DefectType)
    if line:
        query = query.filter(DefectType.line == line)
    return [{"id": dt.id, "name": dt.name, "severity": dt.severity, "line": dt.line} for dt in query.all()]

@app.post("/coils", response_model=CoilOut)
def create_coil(coil: CoilCreate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    if current_user.role == 'inspector':
        raise HTTPException(status_code=403, detail="Not authorized")
    db_coil = Coil(**coil.dict(), created_by=current_user.id)
    db.add(db_coil)
    db.commit()
    db.refresh(db_coil)
    return db_coil

@app.get("/coils", response_model=List[CoilOut])
def read_coils(skip: int = 0, limit: int = 100, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    if current_user.role == 'ua':
        coils = db.query(Coil).offset(skip).limit(limit).all()
    elif current_user.role.endswith('_admin'):
        line = current_user.role.split('_')[0].upper()
        coils = db.query(Coil).filter(Coil.line == line).offset(skip).limit(limit).all()
    else:
        coils = []
    return coils

@app.post("/inspections", response_model=InspectionOut)
def create_inspection(inspection: InspectionCreate, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    if current_user.role == 'inspector':
        raise HTTPException(status_code=403, detail="Not authorized")
    db_inspection = Inspection(
        coil_id=inspection.coil_id,
        inspector_id=inspection.inspector_id,
        inspection_start=inspection.inspection_start,
        inspection_end=inspection.inspection_end,
        fatigue_score_post=inspection.fatigue_score_post,
        missed_defects_estimated=inspection.missed_defects_estimated,
        created_by=current_user.id
    )
    db.add(db_inspection)
    db.flush()
    for d in inspection.defects:
        db_defect = InspectionDefect(
            inspection_id=db_inspection.id,
            defect_type_id=d.defect_type_id,
            position_m=d.position_m,
            quantity=d.quantity
        )
        db.add(db_defect)
    db.commit()
    db.refresh(db_inspection)

    if db_inspection.fatigue_score_post and db_inspection.fatigue_score_post > 8:
        inspector_user = db.query(User).filter(User.inspector_id == db_inspection.inspector_id).first()
        if inspector_user and inspector_user.fcm_token:
            try:
                message = messaging.Message(
                    notification=messaging.Notification(
                        title="Fatigue Alert",
                        body=f"Your fatigue score is {db_inspection.fatigue_score_post}. Take a break!",
                    ),
                    token=inspector_user.fcm_token,
                )
                messaging.send(message)
            except Exception as e:
                print(f"FCM send error: {e}")
    return db_inspection

@app.get("/inspections", response_model=List[InspectionOut])
def read_inspections(skip: int = 0, limit: int = 100, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    if current_user.role == 'ua':
        inspections = db.query(Inspection).offset(skip).limit(limit).all()
    elif current_user.role.endswith('_admin'):
        line = current_user.role.split('_')[0].upper()
        inspections = db.query(Inspection).join(Coil).filter(Coil.line == line).offset(skip).limit(limit).all()
    else:
        inspections = db.query(Inspection).filter(Inspection.inspector_id == current_user.inspector_id).offset(skip).limit(limit).all()
    return inspections

@app.get("/schedules")
def get_schedules(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    today = datetime.now().date()
    if current_user.role == 'ua':
        schedules = db.query(Schedule).filter(Schedule.date >= today).order_by(Schedule.date).all()
    elif current_user.role == 'inspector':
        schedules = db.query(Schedule).filter(
            Schedule.inspector_id == current_user.inspector_id,
            Schedule.date >= today
        ).order_by(Schedule.date).all()
    else:
        schedules = []
    return [{
        "id": s.id,
        "inspector_id": s.inspector_id,
        "date": s.date.isoformat(),
        "shift": s.shift,
        "line": s.line,
        "rotation_flag": s.rotation_flag
    } for s in schedules]

@app.post("/schedules/generate")
def generate_schedules(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != 'ua':
        raise HTTPException(status_code=403, detail="Not authorized")
    start_date = datetime.now().date()
    for day in range(7):
        date = start_date + timedelta(days=day)
        for inspector in db.query(Inspector).all():
            existing = db.query(Schedule).filter(
                Schedule.inspector_id == inspector.inspector_id,
                Schedule.date == date
            ).first()
            if not existing:
                shift = random.choice(["Morning", "Afternoon", "Night"])
                line = random.choice(["CGL", "CAL", "RCL"])
                new_schedule = Schedule(
                    inspector_id=inspector.inspector_id,
                    date=date,
                    shift=shift,
                    line=line,
                    rotation_flag=False
                )
                db.add(new_schedule)
    db.commit()
    return {"message": "Schedules generated for the next 7 days"}

@app.get("/inspector/fatigue_history")
def get_fatigue_history(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != 'inspector':
        raise HTTPException(403, "Not authorized")
    thirty_days_ago = datetime.now() - timedelta(days=30)
    inspections = db.query(Inspection).filter(
        Inspection.inspector_id == current_user.inspector_id,
        Inspection.inspection_start >= thirty_days_ago
    ).order_by(Inspection.inspection_start).all()
    return [{"date": i.inspection_start.isoformat(), "fatigue": i.fatigue_score_post} for i in inspections]

@app.get("/export/pdf")
def export_pdf(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    coils = db.query(Coil).order_by(Coil.created_at.desc()).limit(100).all()
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="TSK Coil Inspection Report", ln=1, align='C')
    pdf.ln(10)
    pdf.set_font("Arial", size=10)
    pdf.cell(40, 10, "Coil ID")
    pdf.cell(40, 10, "Line")
    pdf.cell(40, 10, "Length (m)")
    pdf.cell(40, 10, "Defect Count")
    pdf.ln()
    for c in coils[:20]:
        pdf.cell(40, 10, str(c.coil_id))
        pdf.cell(40, 10, c.line)
        pdf.cell(40, 10, str(c.length_m))
        pdf.cell(40, 10, str(c.defect_count))
        pdf.ln()
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        pdf.output(tmp.name)
        tmp_path = tmp.name
    return FileResponse(tmp_path, filename="TSK_Report.pdf", media_type="application/pdf")

@app.get("/admin/audit_log")
def get_audit_log(skip: int = 0, limit: int = 100, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != 'ua':
        raise HTTPException(status_code=403, detail="Not authorized")
    logs = db.query(AuditLog).order_by(AuditLog.timestamp.desc()).offset(skip).limit(limit).all()
    return [{"id": l.id, "user_id": l.user_id, "action": l.action, "table_name": l.table_name, "record_id": l.record_id, "timestamp": l.timestamp.isoformat()} for l in logs]

@app.get("/api/live/oee")
def get_live_oee(current_user: User = Depends(get_current_user)):
    return {
        "CGL": round(random.uniform(60, 95), 1),
        "CAL": round(random.uniform(65, 98), 1),
        "RCL": round(random.uniform(55, 90), 1),
        "timestamp": datetime.now().isoformat()
    }

# ---------- Analytics endpoints (new) ----------
@app.get("/analytics/defect_pareto")
def defect_pareto(start_date: str, end_date: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Returns top defect types by total quantity within a date range.
    Query parameters: start_date (YYYY-MM-DD), end_date (YYYY-MM-DD)
    """
    start = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
    # Join inspection_defects with inspections and defect_types
    results = (
        db.query(DefectType.name, func.sum(InspectionDefect.quantity).label("total"))
        .join(InspectionDefect, DefectType.id == InspectionDefect.defect_type_id)
        .join(Inspection, Inspection.id == InspectionDefect.inspection_id)
        .filter(Inspection.inspection_start >= start, Inspection.inspection_start < end)
        .group_by(DefectType.name)
        .order_by(func.sum(InspectionDefect.quantity).desc())
        .all()
    )
    return [{"name": r.name, "count": r.total} for r in results]

@app.get("/analytics/line_efficiency")
def line_efficiency(start_date: str, end_date: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Simulated line efficiency (OEE) per line per day.
    For real data, you would have a table of daily OEE. Here we generate random values.
    """
    start = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
    days = (end - start).days
    result = {}
    for line in ["CGL", "CAL", "RCL"]:
        daily = []
        for i in range(days):
            date = start + timedelta(days=i)
            oee = round(random.uniform(65, 95), 1)  # simulated OEE
            daily.append({"date": date.isoformat(), "oee": oee})
        result[line] = daily
    return result

@app.get("/api/a11/predict")
def predict_defect_probability(coil_id: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Predict defect probability for a given coil (dummy implementation).
    In production, replace with a trained ML model.
    """
    coil = db.query(Coil).filter(Coil.coil_id == coil_id).first()
    if not coil:
        raise HTTPException(404, "Coil not found")
    # Dummy: random probability between 0 and 1
    prob = round(random.uniform(0, 1), 2)
    return {"coil_id": coil_id, "defect_probability": prob}

# ---------- Helper for line stats (used in export) ----------
def compute_line_stats(coils, inspections):
    line_stats = {}
    for line in ["CGL", "CAL", "RCL"]:
        line_coils = [c for c in coils if c.line == line]
        n_coils = len(line_coils)
        if n_coils == 0:
            wl_avg = 0.76
            defects_km_avg = 0.5
            cv = 0.3
            fat_avg = 6.5
        else:
            wls = []
            defects_km = []
            for c in line_coils:
                if c.length_m and c.speed_mps:
                    dur_min = c.length_m / (c.speed_mps * 60)
                    wl = (dur_min * 60) / c.length_m if c.length_m else 0
                    wls.append(wl)
                if c.length_m:
                    defects_km.append(c.defect_count / (c.length_m / 1000))
            wl_avg = np.mean(wls) if wls else 0.76
            defects_km_avg = np.mean(defects_km) if defects_km else 0.5
            cv = np.std(defects_km) / defects_km_avg if defects_km_avg > 0 and len(defects_km) > 1 else 0.3
            line_inspections = [i for i in inspections if i.coil_id in [c.id for c in line_coils]]
            fats = [i.fatigue_score_post for i in line_inspections if i.fatigue_score_post]
            fat_avg = np.mean(fats) if fats else 6.5
        speed = {"CGL": 150, "CAL": 200, "RCL": 250}[line]
        raw = (speed * wl_avg) / 60
        n_base = max(1, math.ceil(raw))
        n_peak = math.ceil(n_base * 1.3)
        line_stats[line] = {
            "wl_avg": round(wl_avg, 4),
            "defects_km_avg": round(defects_km_avg, 4),
            "n_coils": n_coils,
            "speed": speed,
            "n_base": n_base,
            "n_peak": n_peak,
            "cv": round(cv, 4),
            "fat_avg": round(fat_avg, 2)
        }
    return line_stats

@app.get("/export")
def export_results(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role == 'ua':
        coils = db.query(Coil).order_by(Coil.created_at.desc()).all()
        inspections = db.query(Inspection).all()
        inspectors = db.query(Inspector).all()
    elif current_user.role.endswith('_admin'):
        line = current_user.role.split('_')[0].upper()
        coils = db.query(Coil).filter(Coil.line == line).order_by(Coil.created_at.desc()).all()
        inspections = db.query(Inspection).join(Coil).filter(Coil.line == line).all()
        inspectors = db.query(Inspector).all()
    else:
        coils = []
        inspections = db.query(Inspection).filter(Inspection.inspector_id == current_user.inspector_id).all()
        inspectors = db.query(Inspector).filter(Inspector.inspector_id == current_user.inspector_id).all()

    coils_df = pd.DataFrame([c.__dict__ for c in coils]) if coils else pd.DataFrame()
    inspections_df = pd.DataFrame([i.__dict__ for i in inspections]) if inspections else pd.DataFrame()
    inspectors_df = pd.DataFrame([ins.__dict__ for ins in inspectors]) if inspectors else pd.DataFrame()
    line_stats = compute_line_stats(coils, inspections)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    build_dashboard(wb, coils, inspections, inspectors, line_stats)
    build_coil_detail(wb, coils, inspections)
    build_fatigue_log(wb, inspections, coils)
    build_inspector_matrix(wb, inspectors)
    build_inspector_calc(wb, line_stats)
    build_change_log(wb, len(coils))

    try:
        from ai_engine import add_ai_sheets_to_workbook
        wb = add_ai_sheets_to_workbook(wb, coils_df, inspections_df, inspectors_df, line_stats, avail_insp=len(inspectors))
    except Exception as e:
        err_ws = wb.create_sheet("AI Error")
        err_ws.append(["Error", str(e)])
        print(f"AI sheets error: {e}")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    return FileResponse(
        path=tmp_path,
        filename=f"TSK_Final_Results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/api/a1/forecast")
def get_a1_forecast(current_user: User = Depends(get_current_user)):
    forecast = [{"hour": i, "defects": round(i * 0.5, 2)} for i in range(1, 25)]
    return forecast

@app.get("/api/a1/forecast/line")
def get_a1_forecast_by_line(line: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != 'ua' and current_user.line != line:
        raise HTTPException(status_code=403, detail="Not authorized for this line")
    coils = db.query(Coil).filter(Coil.line == line).all()
    if not coils:
        defects_km = [0.5] * 24
    else:
        avg_def_km = np.mean([c.defect_count / (c.length_m/1000) for c in coils if c.length_m]) or 0.5
        defects_km = [round(avg_def_km * (1 + 0.02 * i), 4) for i in range(24)]
    forecast = [{"hour": i+1, "defects": defects_km[i]} for i in range(24)]
    return forecast

@app.get("/api/a4/fatigue_predict")
def get_a4_fatigue_predict(current_user: User = Depends(get_current_user)):
    pred = [{"hour": i, "fatigue": round(5 + i * 0.2, 1)} for i in range(1, 13)]
    return pred

@app.get("/api/a4/fatigue_predict/inspector")
def get_fatigue_for_inspector(inspector_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != 'ua' and current_user.inspector_id != inspector_id:
        raise HTTPException(status_code=403, detail="Not authorized")
    inspections = db.query(Inspection).filter(Inspection.inspector_id == inspector_id).order_by(Inspection.inspection_start).all()
    if not inspections:
        pred = [{"hour": i+1, "fatigue": round(5 + i * 0.2, 1)} for i in range(12)]
    else:
        last_fatigue = inspections[-1].fatigue_score_post if inspections[-1].fatigue_score_post else 5
        pred = [{"hour": i+1, "fatigue": round(min(10, last_fatigue + i * 0.1), 1)} for i in range(12)]
    return pred

@app.get("/admin/users")
def list_users(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != 'ua':
        raise HTTPException(status_code=403, detail="Not authorized")
    users = db.query(User).all()
    return [{"id": u.id, "email": u.email, "role": u.role, "line": u.line, "inspector_id": u.inspector_id, "name": u.name} for u in users]

@app.delete("/admin/users/{user_id}")
def delete_user(user_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != 'ua':
        raise HTTPException(status_code=403, detail="Not authorized")
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    db.delete(user)
    db.commit()
    return {"message": "User deleted"}

@app.get("/")
def root():
    return {"message": "TSK Backend is alive. Use /docs for API documentation."}